"""
DE 5주차 채점 결과 엑셀 생성 스크립트
- 0 or 10점 이진 채점 (부분점수 없음)
- 10문제 x 10점 = 100점 만점
- 출력: /home/pc/dev/metacode/MetacodeAssessment/DE/5week/채점결과.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "DE 5주차 채점"

# 학생 13명 (가나다 순)
students = [
    "김나경", "김민성", "김민주", "김민혁", "김윤재",
    "신민재", "안주영", "오창석", "이민호", "임석제",
    "장재훈", "정찬엽", "주세용",
]

# 행별 데이터: 헤더 + Q1~Q10 + 총점
rows = [
    ["문제"] + students,
    ["Q1",
     "O", "O", "O", "O", "O",
     "O", "O", "O", "O", "O",
     "O", "O", "O"],
    ["Q2",
     "O", "O", "O", "O", "O",
     "O", "O", "O", "O", "O",
     "O", "Databricks Volumes 사용(S3/s3a:// 미사용)", "O"],
    ["Q3",
     "O", "O", "O", "O", "O",
     "O", "O", "O", "O", "O",
     "O", "Databricks Volumes 사용(S3/s3a:// 미사용)", "O"],
    ["Q4",
     "O", "O", "O", "O", "O",
     "O", "O", "O", "O", "O",
     "O", "O", "O"],
    ["Q5",
     "O", "O", "O", "O", "O",
     "O", "O", "O", "O", "S3 미사용 - inline 리스트로 DataFrame 생성",
     "O", "Databricks Volumes 사용(S3/s3a:// 미사용)", "O"],
    ["Q6",
     "O", "O", "O", "O", ".pages 열람 불가 - Dataset 필터 구현 확인 불가",
     "O", "O", "O", "O", "O",
     "O", "O", "O"],
    ["Q7",
     "O", "O", "O", "O", ".pages 열람 불가 - UDF 구현 확인 불가",
     "O", "O", "get_user_segment 함수 정의 누락 (udf 호출만 존재)", "O", "O",
     "O", "O", "O"],
    ["Q8",
     "PDF 내 Airflow DAG 코드 없음",
     "Airflow DAG 파일 미제출 (spark_user_order_dag.py 없음)",
     "O", "O", ".pages 열람 불가 - DAG 구현 확인 불가",
     "O", "O", "O", "O", "O",
     "DAG 섹션에 \"에어플로우:\" 제목만 존재, 실제 DAG 코드 없음",
     "O", "Airflow DAG 코드 미제출"],
    ["Q9",
     "O", "O", "users.csv, transactions.csv S3 업로드 캡쳐 없음", "O", "O",
     "O", "O", "O", "O", "O",
     "O", "Databricks Volumes 사용(S3/s3a:// 미사용)", "O"],
    ["Q10",
     "O", "O", "GROUP BY 사용자별 집계 없음 (SUM/COUNT 집계 함수 미사용)", "O", "O",
     "O", "O", "O", "O", "O",
     "O", "O", "O"],
    ["총점", 90, 90, 80, 100, 70, 100, 100, 90, 100, 90, 90, 60, 90],
]

# 복붙용 피드백 (학생별 총점 + 틀린 문제번호/사유)
feedback = {
    "김나경": "90\n8. PDF 내 Airflow DAG 코드 없음",
    "김민성": "90\n8. Airflow DAG 파일 미제출 (spark_user_order_dag.py 없음)",
    "김민주": "80\n9. users.csv, transactions.csv S3 업로드 캡쳐 없음\n10. Spark SQL 쿼리에 GROUP BY 사용자별 집계 없음 (문제지 요구: 사용자별 총 거래금액 또는 거래 건수 집계)",
    "김민혁": "100\n(감점 없음)",
    "김윤재": "70\n6. .pages 열람 불가 - Dataset 필터 구현 확인 불가\n7. .pages 열람 불가 - UDF 구현 확인 불가\n8. .pages 열람 불가 - DAG 구현 확인 불가",
    "신민재": "100\n(감점 없음)",
    "안주영": "100\n(감점 없음)",
    "오창석": "90\n7. get_user_segment 함수 정의 누락 (udf 호출만 존재)",
    "이민호": "100\n(감점 없음)",
    "임석제": "90\n5. S3 미사용 - inline 리스트로 DataFrame 생성",
    "장재훈": "90\n8. DAG 섹션에 \"에어플로우:\" 제목만 존재, 실제 DAG 코드 없음",
    "정찬엽": "60\n2. Databricks Volumes 사용 (S3/s3a:// 미사용)\n3. Databricks Volumes 사용 (S3/s3a:// 미사용)\n5. Databricks Volumes 사용 (S3/s3a:// 미사용)\n9. Databricks Volumes 사용 (S3/s3a:// 미사용)",
    "주세용": "90\n8. Airflow DAG 코드 미제출",
}
rows.append(["복붙용 피드백"] + [feedback[s] for s in students])

for r in rows:
    ws.append(r)

# 스타일 정의
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
qcol_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
qcol_font = Font(name="맑은 고딕", bold=True, size=11)
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ok_font = Font(name="맑은 고딕", color="006100", bold=True, size=11)
bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
bad_font = Font(name="맑은 고딕", color="9C0006", size=10)
empty_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
empty_font = Font(name="맑은 고딕", color="9C5700", bold=True, size=10)
total_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
total_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=12)
feedback_label_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
feedback_label_font = Font(name="맑은 고딕", bold=True, size=11, color="000000")
feedback_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
feedback_font = Font(name="맑은 고딕", size=10, color="000000")

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

max_col = ws.max_column

# 헤더 행 (row 1)
for c in range(1, max_col + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# Q1~Q10 행 (row 2~11)
for r in range(2, 12):
    qcell = ws.cell(row=r, column=1)
    qcell.fill = qcol_fill
    qcell.font = qcol_font
    qcell.alignment = center
    qcell.border = border
    for c in range(2, max_col + 1):
        cell = ws.cell(row=r, column=c)
        v = str(cell.value)
        if v == "O":
            cell.fill = ok_fill
            cell.font = ok_font
            cell.alignment = center
        elif v in ("미작성", "미제출"):
            cell.fill = empty_fill
            cell.font = empty_font
            cell.alignment = center
        else:
            cell.fill = bad_fill
            cell.font = bad_font
            cell.alignment = left
        cell.border = border

# 총점 행 (row 12)
for c in range(1, max_col + 1):
    cell = ws.cell(row=12, column=c)
    cell.fill = total_fill
    cell.font = total_font
    cell.alignment = center
    cell.border = border

# 복붙용 피드백 행 (row 13) — 80점 이하는 연한 빨강 4
low_score_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
totals_row = rows[11][1:]
for c in range(1, max_col + 1):
    cell = ws.cell(row=13, column=c)
    cell.border = border
    if c == 1:
        cell.fill = feedback_label_fill
        cell.font = feedback_label_font
        cell.alignment = center
    else:
        student_total = totals_row[c - 2]
        if student_total <= 80:
            cell.fill = low_score_fill
        else:
            cell.fill = feedback_fill
        cell.font = feedback_font
        cell.alignment = top_left

# 열 너비
ws.column_dimensions['A'].width = 14
for c in range(2, max_col + 1):
    ws.column_dimensions[get_column_letter(c)].width = 28

# 행 높이
ws.row_dimensions[1].height = 26
for r in range(2, 12):
    ws.row_dimensions[r].height = 48
ws.row_dimensions[12].height = 26
ws.row_dimensions[13].height = 220

ws.freeze_panes = "B2"

out = "/home/pc/dev/metacode/MetacodeAssessment/DE/5week/채점결과.xlsx"
wb.save(out)
print("saved:", out)
