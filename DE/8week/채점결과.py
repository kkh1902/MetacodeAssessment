import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "8주차 채점"

# 스타일 정의
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
task_header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
total_font = Font(bold=True, size=12)
name_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
name_font = Font(bold=True, size=11)

# 채점 방법
# - S3 관련 과제(과제2, 과제9): 실제 AWS S3에 접속할 수 없으므로
#   LocalStack(가짜 S3)을 띄워 학생 코드를 실행하여 동작 여부 확인
# - 그 외 과제: 코드 리뷰 + 스크린샷 확인

# 과제 항목 (10개 × 10점 = 100점)
tasks = [
    ("1-1", "hello_airflow_dag", "기본 DAG 작성 (BashOperator)"),
    ("1-2", "s3_download_dag", "S3 파일 다운로드 DAG"),
    ("2-1", "analyze_csv.py", "PySpark CSV 분석 스크립트"),
    ("2-2", "run_spark_job.py", "Airflow-Spark 연동"),
    ("3", "db_query_dag", "PostgreSQL 조회 DAG"),
    ("4", "apply_udf.py", "PySpark UDF 적용"),
    ("5", "stream_log_processor.py", "Spark Streaming 구현"),
    ("6", "pipeline_dag", "통합 파이프라인 DAG"),
    ("7/8", "fetch_subway_data", "지하철 데이터 수집 + PySpark 분석"),
    ("9/10", "Kubernetes", "nginx 배포 + HPA 설정"),
]

# 채점 데이터: "O" = 10점, 그 외 문자열 = X (0점) + 사유
# 문제 9: 수집(Task1) + 분석(Task2) 모두 충족해야 O
# 문제 10: Deployment+Service + HPA 모두 충족해야 O
# fmt: off
scores = {
    "강경래": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - UDF 적용 후 결과 저장(write) 누락, show()만 수행",
        "O",
        "O",
        "O",
        "O",
    ],
    "김건": [
        "X - hello_airflow_dag.py 미제출",
        "X - Airflow Connection 미활용 (boto3에 키 직접 입력)",
        "O",
        "X - spark-submit 미사용 (subprocess.run으로 직접 실행)",
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - HPA 미제출, 동작 결과 없음",
    ],
    "김민지": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
    ],
    "김준규": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
    ],
    "김혜령": [
        "O",
        "O",
        "O",
        "X - --master 옵션 미지정",
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - HPA 동작 결과 스크린샷/YAML 없음",
    ],
    "남덕우": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - UDF 적용 후 결과 저장(write) 누락, show()만 수행",
        "O",
        "O",
        "O",
        "O",
    ],
    "도건호": [
        "O",
        "O",
        "O",
        "X - spark-submit 미사용 (PythonOperator로 직접 실행)",
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - HPA YAML/스크린샷 미확인",
    ],
    "박찬익": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
    ],
    "배상훈": [
        "O",
        "O",
        "O",
        "X - spark-submit 미사용 (PythonOperator로 직접 실행)",
        "O",
        "X - UDF 적용 후 결과 저장(write) 누락, show()만 수행",
        "O",
        "O",
        "O",
        "X - HPA YAML/스크린샷 미제출",
    ],
    "배형진": [
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "O",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "O",
        "X - 실행 스크린샷/로그 미제출",
    ],
    "원승환": [
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "O",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "X - 실행 스크린샷/로그 미제출",
        "O",
        "X - 실행 스크린샷/로그 미제출",
    ],
    "이범학": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
    ],
    "이상운": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - 실행 증빙(스크린샷) 없음",
        "X - K8s YAML/스크린샷 미제출",
    ],
    "이소래": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - 파일 미제출",
        "X - K8s 파일/스크린샷 미제출",
    ],
    "홍세현": [
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
    ],
    "황하윤": [
        "O",
        "O",
        "X - analyze_csv.py 미제출",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - Service 없음, HPA 없음",
    ],
    "박지민": [
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
        "X - 코드가 이미지(스크린샷)로만 제출, 텍스트 코드 없음",
    ],
    "우수빈": [
        "O",
        "O",
        "O",
        "O",
        "X - psycopg2 미사용, 실제 쿼리 미수행(print만)",
        "O",
        "O",
        "O",
        "O",
        "X - K8s YAML/스크린샷 미제출",
    ],
    "백경희": [
        "O",
        "X - S3 미사용(shutil 로컬 복사만), BaseHook/S3Hook 미사용",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "O",
        "X - K8s YAML/스크린샷 미제출",
    ],
}
# fmt: on

# ── 레이아웃 ──
num_tasks = len(tasks)
headers = ["이름"] + [f"과제{t[0]}" for t in tasks] + ["총점"]
col_count = len(headers)

# 헤더 작성
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill if col_idx == 1 or col_idx == col_count else task_header_fill
    cell.alignment = center
    cell.border = thin_border

# 열 너비 설정
ws.column_dimensions["A"].width = 12  # 이름
for i in range(2, num_tasks + 2):
    ws.column_dimensions[get_column_letter(i)].width = 28
ws.column_dimensions[get_column_letter(col_count)].width = 10  # 총점

# 학생 데이터 행 (가나다순 정렬됨)
for row_idx, (name, task_scores) in enumerate(scores.items(), 2):
    # 이름
    name_cell = ws.cell(row=row_idx, column=1, value=name)
    name_cell.font = name_font
    name_cell.fill = name_fill
    name_cell.alignment = center
    name_cell.border = thin_border

    total = 0
    for col_idx, score in enumerate(task_scores, 2):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border

        if score == "O":
            cell.value = "O"
            cell.fill = pass_fill
            cell.alignment = center
            cell.font = Font(bold=True, size=11, color="006100")
            total += 10
        else:
            cell.value = score
            cell.fill = fail_fill
            cell.alignment = left_wrap
            cell.font = Font(size=9, color="9C0006")

    # 총점
    total_cell = ws.cell(row=row_idx, column=col_count, value=total)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = center
    total_cell.border = thin_border

# 행 높이 설정
ws.row_dimensions[1].height = 30
for row_idx in range(2, len(scores) + 2):
    ws.row_dimensions[row_idx].height = 50

# 오답 정리 열 추가 (총점 오른쪽)
reason_col = col_count + 1
ws.column_dimensions[get_column_letter(reason_col)].width = 60

# 헤더
reason_header = ws.cell(row=1, column=reason_col, value="오답 정리")
reason_header.font = header_font
reason_header.fill = header_fill
reason_header.alignment = center

for row_idx, (name, task_scores) in enumerate(scores.items(), 2):
    x_list = []
    for i, score in enumerate(task_scores):
        if score != "O":
            task_no = tasks[i][0]
            reason = score.replace("X - ", "").replace("X -", "").strip()
            x_list.append(f"{task_no} ({reason})")

    total = sum(10 for s in task_scores if s == "O")
    if x_list:
        reason_text = f"{total}\n" + "\n".join(x_list)
    else:
        reason_text = f"{total}\n만점"

    cell = ws.cell(row=row_idx, column=reason_col, value=reason_text)
    cell.font = Font(size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if not x_list:
        cell.font = Font(bold=True, size=11, color="006100")

    # 오답 많으면 행 높이 늘리기
    ws.row_dimensions[row_idx].height = max(50, len(x_list) * 18)

# 틀 고정 (이름 열 + 헤더 행)
ws.freeze_panes = "B2"

output_path = "/home/pc/dev/metade/8week/8주차_채점결과.xlsx"
wb.save(output_path)
print(f"저장 완료: {output_path}")
print(f"학생 수: {len(scores)}명")
print()
for name, task_scores in scores.items():
    total = sum(10 for s in task_scores if s == "O")
    print(f"  {name}: {total}/100점")
